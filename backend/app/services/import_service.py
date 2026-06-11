import io
import re
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.models.geographic import Department, Province, District
from app.models.master_data import DocumentType, BusinessType, ClientGroup
from app.models.client import Client
from app.models.user import User
from app.schemas.client import ClientCreate
from app.crud import crud_client

# Columnas obligatorias en snake_case normalizado
REQUIRED_COLUMNS = {
    "dia_de_visita",
    "cod_vendedor",
    "vendedor",
    "cod_cliente",
    "razon_social",
    "tipo_de_documento",
    "n_documuento",
    "direccion",
    "distrito",
    "provincia",
    "departamento",
    "tipo_de_negocio",
    "grupo_cliente",
    "celular",
}


def _normalize_header(header: str) -> str:
    """
    Normaliza los encabezados del Excel: quita tildes, caracteres especiales,
    reemplaza espacios por subguiones y lo pasa a minúsculas (snake_case).
    """
    if not header:
        return ""
    replacements = {'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ñ': 'N'}
    s = header.strip().upper()
    for k, v in replacements.items():
        s = s.replace(k, v)
    
    s = s.replace("N°", "N").replace("N° ", "N_")
    s = re.sub(r'[^A-Z0-9\s_]', '', s)
    s = re.sub(r'\s+', '_', s)
    return s.lower()


def _build_lookup_cache(db: Session) -> Dict[str, Any]:
    """Pre-carga todos los datos de referencia en memoria para evitar N+1 queries."""
    departments = db.query(Department).all()
    provinces = db.query(Province).all()
    districts = db.query(District).all()
    doc_types = db.query(DocumentType).all()
    business_types = db.query(BusinessType).all()
    client_groups = db.query(ClientGroup).all()

    # Pre-cargar usuarios indexados por código para búsqueda O(1)
    users = db.query(User).all()

    # Pre-cargar códigos de clientes existentes para validar duplicados
    existing_codes = {c[0].upper() for c in db.query(Client.code).filter(Client.code.isnot(None)).all()}

    return {
        "departments": {d.name.upper(): d for d in departments},
        "provinces": {p.name.upper(): p for p in provinces},
        "districts": {dist.name.upper(): dist for dist in districts},
        "doc_types": {dt.description.upper(): dt for dt in doc_types},
        "business_types": {bt.description.upper(): bt for bt in business_types},
        "client_groups": {cg.description.upper(): cg for cg in client_groups},
        "users_by_code": {u.code.upper(): u for u in users if u.code},
        "existing_client_codes": existing_codes,
    }


def _resolve_district_id(
    cache: Dict[str, Any],
    departamento: str,
    provincia: str,
    distrito: str,
) -> Optional[int]:
    """Busca el district_id de forma jerárquica. Retorna None si no existe."""
    if not departamento or not provincia or not distrito:
        return None
    try:
        dep = cache["departments"].get(departamento.upper())
        if dep is None:
            return None

        prov = cache["provinces"].get(provincia.upper())
        if prov is None or prov.department_id != dep.id:
            return None

        dist = cache["districts"].get(distrito.upper())
        if dist is None or dist.province_id != prov.id:
            return None

        return dist.id
    except Exception:
        return None


def _to_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _to_float(val) -> Optional[float]:
    try:
        return float(val) if val not in (None, "") else None
    except (ValueError, TypeError):
        return None


def process_excel_import(
    file_bytes: bytes,
    db: Session,
) -> Dict[str, Any]:
    """
    Procesa el archivo Excel. El user_id de cada cliente se determina a partir de
    cod_vendedor en cada fila (no del usuario autenticado).

    Si cod_vendedor está vacío o no existe en BD → aborta toda la importación.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValueError("El archivo no es un Excel válido (.xlsx).")

    ws = wb.active

    # Leer y normalizar encabezados a snake_case
    raw_headers = [_to_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [_normalize_header(h) for h in raw_headers]
    header_index = {h: i for i, h in enumerate(headers)}

    # Validar columnas obligatorias
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        friendly_missing = [m.replace("_", " ").upper() for m in missing]
        raise ValueError(f"Campos faltantes en el Excel: {', '.join(sorted(friendly_missing))}")

    # Pre-cargar cache de lookups
    cache = _build_lookup_cache(db)

    errors: List[Dict[str, Any]] = []
    created = 0
    omitted = 0

    def get_cell(row_values: list, col_name: str) -> str:
        idx = header_index.get(col_name)
        if idx is None or idx >= len(row_values):
            return ""
        return _to_str(row_values[idx])

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue  # Fila vacía

        row_vals = list(row)

        try:
            # ── Resolución de vendedor ──────────────────────────────────────
            cod_vendedor = get_cell(row_vals, "cod_vendedor").strip().upper()

            if not cod_vendedor:
                raise ValueError(
                    "El campo 'cod_vendedor' es obligatorio y está vacío en esta fila."
                )

            seller = cache["users_by_code"].get(cod_vendedor)
            if seller is None:
                raise ValueError(
                    f"No existe ningún usuario con el código de vendedor '{cod_vendedor}'."
                )

            assigned_user_id = seller.id

            # ── Extracción y normalización de campos del cliente ────────────
            codigo_cliente     = get_cell(row_vals, "cod_cliente").upper()
            nombre             = get_cell(row_vals, "razon_social").upper()
            tipo_documento_txt = get_cell(row_vals, "tipo_de_documento").upper()
            num_documento      = get_cell(row_vals, "n_documuento")
            direccion          = get_cell(row_vals, "direccion").upper()
            distrito_txt       = get_cell(row_vals, "distrito").upper()
            provincia_txt      = get_cell(row_vals, "provincia").upper()
            departamento_txt   = get_cell(row_vals, "departamento").upper()
            tipo_negocio_txt   = get_cell(row_vals, "tipo_de_negocio").upper()
            grupo_cliente_txt  = get_cell(row_vals, "grupo_cliente").upper()
            telefono           = get_cell(row_vals, "telefono") or None
            celular            = get_cell(row_vals, "celular") or None
            latitud            = _to_float(get_cell(row_vals, "latitud"))
            longitud           = _to_float(get_cell(row_vals, "longitud"))
            observacion        = get_cell(row_vals, "observacion").upper()

            # ── Validar duplicado por código ────────────────────────────────
            if codigo_cliente and codigo_cliente in cache["existing_client_codes"]:
                omitted += 1
                errors.append({
                    "fila": row_num,
                    "error": f"Cliente con código '{codigo_cliente}' ya existe en el sistema (omitido)."
                })
                continue

            # ── Validar duplicado por número de documento ───────────────────
            if num_documento:
                existing_doc = db.query(Client).filter(Client.document_number == num_documento).first()
                if existing_doc:
                    omitted += 1
                    errors.append({
                        "fila": row_num,
                        "error": f"Cliente con documento '{num_documento}' ya existe (omitido)."
                    })
                    continue

            # ── Resolver IDs desde textos ───────────────────────────────────
            doc_type        = cache["doc_types"].get(tipo_documento_txt) if tipo_documento_txt else None
            doc_type_id     = doc_type.id if doc_type else None

            business_type   = cache["business_types"].get(tipo_negocio_txt) if tipo_negocio_txt else None
            business_type_id = business_type.id if business_type else None

            client_group    = cache["client_groups"].get(grupo_cliente_txt) if grupo_cliente_txt else None
            client_group_id = client_group.id if client_group else None

            district_id = None
            if departamento_txt or provincia_txt or distrito_txt:
                district_id = _resolve_district_id(
                    cache, departamento_txt, provincia_txt, distrito_txt
                )

            # ── Construir y persistir el cliente ────────────────────────────
            client_data = ClientCreate(
                code=codigo_cliente[:6] if codigo_cliente else None,
                name=nombre[:50] if nombre else None,
                document_type_id=doc_type_id,
                document_number=num_documento[:11] if num_documento else None,
                address=direccion if direccion else None,
                district_id=district_id,
                business_type_id=business_type_id,
                client_group_id=client_group_id,
                cellphone=celular[:9] if celular else None,
                telephone=telefono[:9] if telefono else None,
                active=True,
                user_id=assigned_user_id,
                latitud=latitud,
                longitud=longitud,
                observation=observacion if observacion else "SIN OBSERVACIÓN",
            )

            crud_client.create_client(db, client_in=client_data)

            # Actualizar caché de códigos para detectar duplicados dentro del mismo archivo
            if codigo_cliente:
                cache["existing_client_codes"].add(codigo_cliente)

            created += 1

        except Exception as e:
            errors.append({"fila": row_num, "error": str(e)})
            omitted += 1
            db.rollback()

    wb.close()

    total = created + omitted
    return {
        "total_registros": total,
        "creados": created,
        "omitidos": omitted,
        "errores": errors,
    }